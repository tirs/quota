"""
Export utilities for PDF, Excel, CSV and other formats
"""
import pandas as pd
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import Database

db = Database()

def export_quotes_to_excel(quotes_data: list, filename: str = "quotes") -> BytesIO:
    """Export quotes to Excel with formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotes"
    
    # Define styles
    header_fill = PatternFill(start_color="00D9FF", end_color="00D9FF", fill_type="solid")
    header_font = Font(bold=True, color="161B22", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Quote #", "Customer", "Email", "Status", "Subtotal", "Tax", "Total", "Created", "Updated"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data
    for quote_id in quotes_data:
        quote = db.get_quote(quote_id) if isinstance(quote_id, int) else quote_id
        if quote:
            customer = db.get_customer_by_id(quote['customer_id'])
            row = [
                quote['quote_number'],
                customer['name'] if customer else "",
                customer['email'] if customer else "",
                quote['status'].upper(),
                quote['subtotal'],
                quote['tax_amount'],
                quote['total'],
                quote['created_at'],
                quote['updated_at']
            ]
            ws.append(row)
    
    # Set column widths
    column_widths = [15, 20, 25, 12, 12, 12, 12, 20, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Format numbers
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=7):
        for cell in row:
            cell.number_format = '$#,##0.00'
            cell.border = border
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def export_quotes_to_detailed_excel(quote_ids: list) -> BytesIO:
    """Export quotes with detailed line items to Excel"""
    wb = Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="00D9FF", end_color="00D9FF", fill_type="solid")
    header_font = Font(bold=True, color="161B22", size=12)
    title_font = Font(bold=True, size=14, color="00D9FF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for idx, quote_id in enumerate(quote_ids):
        quote = db.get_quote(quote_id)
        if not quote:
            continue
        
        customer = db.get_customer_by_id(quote['customer_id'])
        items = db.get_quote_items(quote_id)
        
        # Create sheet
        ws = wb.create_sheet(f"Quote_{quote['quote_number'].split('-')[-1]}")
        
        # Title
        ws['A1'] = f"Quote: {quote['quote_number']}"
        ws['A1'].font = title_font
        ws.merge_cells('A1:E1')
        
        # Customer info
        row = 3
        ws[f'A{row}'] = "Customer:"
        ws[f'B{row}'] = customer['name'] if customer else ""
        row += 1
        ws[f'A{row}'] = "Email:"
        ws[f'B{row}'] = customer['email'] if customer else ""
        row += 1
        ws[f'A{row}'] = "Status:"
        ws[f'B{row}'] = quote['status'].upper()
        
        # Line items header
        row = 7
        item_headers = ["Product", "Quantity", "Unit Price", "Line Total"]
        for col, header in enumerate(item_headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        # Line items
        row = 8
        for item in items:
            ws.cell(row=row, column=1).value = item['name']
            ws.cell(row=row, column=2).value = item['quantity']
            ws.cell(row=row, column=3).value = item['unit_price']
            ws.cell(row=row, column=4).value = item['line_total']
            row += 1
        
        # Totals
        row += 1
        ws.cell(row=row, column=3).value = "Subtotal:"
        ws.cell(row=row, column=4).value = quote['subtotal']
        ws.cell(row=row, column=4).number_format = '$#,##0.00'
        
        row += 1
        ws.cell(row=row, column=3).value = "Tax:"
        ws.cell(row=row, column=4).value = quote['tax_amount']
        ws.cell(row=row, column=4).number_format = '$#,##0.00'
        
        row += 1
        ws.cell(row=row, column=3).value = "TOTAL:"
        total_cell = ws.cell(row=row, column=4)
        total_cell.value = quote['total']
        total_cell.number_format = '$#,##0.00'
        total_cell.font = Font(bold=True, size=12)
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    # Remove default sheet if it exists and we have real sheets
    if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb['Sheet']
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def export_analytics_report_to_excel(intelligence_data: dict) -> BytesIO:
    """Export analytics report to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Analytics"
    
    header_fill = PatternFill(start_color="3FB950", end_color="3FB950", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Key metrics
    row = 1
    ws[f'A{row}'] = "KEY METRICS"
    ws[f'A{row}'].font = Font(bold=True, size=14, color="3FB950")
    
    metrics = [
        ("Total Customers", intelligence_data.get('total_customers', 0)),
        ("Total Quotes", intelligence_data.get('total_quotes', 0)),
        ("Total Value", intelligence_data.get('total_value', 0)),
        ("Win Rate (%)", intelligence_data.get('win_rate', 0)),
        ("Average Deal Size", intelligence_data.get('average_deal_size', 0)),
        ("30-Day Revenue", intelligence_data.get('recent_30_day_value', 0)),
    ]
    
    row = 3
    for metric_name, value in metrics:
        ws[f'A{row}'] = metric_name
        ws[f'B{row}'] = value if not isinstance(value, float) else f"${value:,.2f}"
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1
    
    # Forecast
    row += 2
    ws[f'A{row}'] = "REVENUE FORECAST (30 Days)"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="3FB950")
    
    forecast = intelligence_data.get('forecast', {})
    row += 1
    ws[f'A{row}'] = "Forecasted Revenue"
    ws[f'B{row}'] = f"${forecast.get('forecast', 0):,.2f}"
    
    row += 1
    ws[f'A{row}'] = "Daily Average"
    ws[f'B{row}'] = f"${forecast.get('daily_average', 0):,.2f}"
    
    row += 1
    ws[f'A{row}'] = "Confidence"
    ws[f'B{row}'] = forecast.get('confidence', 'Unknown')
    
    row += 1
    ws[f'A{row}'] = "Trend"
    ws[f'B{row}'] = forecast.get('trend', 'Unknown')
    
    # Set widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def export_customer_health_report() -> BytesIO:
    """Export customer health scores report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Health"
    
    header_fill = PatternFill(start_color="FF006E", end_color="FF006E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Customer", "Engagement Score", "Spend Score", "Growth Score", "Health Score", "Risk Level"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    health_scores = db.get_all_customer_health_scores()
    for row, score in enumerate(health_scores, 2):
        ws.cell(row=row, column=1).value = score['name']
        ws.cell(row=row, column=2).value = round(score['engagement_score'], 1)
        ws.cell(row=row, column=3).value = round(score['spend_score'], 1)
        ws.cell(row=row, column=4).value = round(score['growth_score'], 1)
        ws.cell(row=row, column=5).value = round(score['health_score'], 1)
        ws.cell(row=row, column=6).value = score['risk_level']
        
        # Color code risk level
        risk_cell = ws.cell(row=row, column=6)
        if score['risk_level'] == 'HIGH':
            risk_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        elif score['risk_level'] == 'MEDIUM':
            risk_cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
        else:
            risk_cell.fill = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def export_audit_log_to_csv() -> BytesIO:
    """Export audit log to CSV"""
    logs = db.get_audit_logs(limit=1000)
    
    df = pd.DataFrame(logs)
    if len(df) > 0:
        df = df[['user', 'action', 'entity_type', 'entity_id', 'details', 'created_at']]
        df.columns = ['User', 'Action', 'Entity Type', 'Entity ID', 'Details', 'Created At']
    
    buffer = BytesIO()
    df.to_csv(buffer, index=False)
    buffer.seek(0)
    return buffer